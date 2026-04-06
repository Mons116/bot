import asyncio
import logging
import os
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import Message, BufferedInputFile

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

TOKEN = os.getenv("BOT_TOKEN")

logging.basicConfig(level=logging.INFO)

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# ====================== ГРУППЫ ======================

groups = {
    "Пальто": ["drap_2-beige","drap_2-black","drap_2-grey","drap-beige","drap-brown","drap-grey","drap-black","coat-jacket-beige","coat-jacket-grey","coat-jacket-brown","drap_3(belt)-grey","drap_3(belt)-brown","drap_3(belt)-grafit","drap_3(belt)-beige"],
    "Куртки": ["suede-ohra","suede-milk","suede-brown","bomber-ohra","bomber-brown","bomber-milk"],
    "двойка топ с завязками": ["bows-blue","bows-cappuccino","bows-haki"],
    "Песок": ["costum-black","costum-blue","costum-brown","costum-green","costum-grey","costum-olive"],
    "Шанель": ["flax-beige","flax-blue","flax-brown","flax-green","flax-grey"],
    "Шорты": ["short1beige","short1blue","short1fuksia","short1green","short-haki","short-brown","short-black","short-mentol23"],
    "Комбез": ["office-grey","office-dark blue", "office-beige", "plecho-pants-white", "plecho-pants-yellow", "plecho-pants-dark blue", "plecho-pants-bork", "komb(на запах)- beige", "komb(на запах)-dark blue", "komb(на запах)-bork"],
    "Зара": ["zara-blue","zara-brown","zara-beige"],
    "франц.лен": ["french-mocha","french-dark blue","french-blue", "french-bork"],
    "платья": ["dress-yellow", "dress-black", "dress-blue", "dress-white"]
}

group_colors = {
    "Пальто": "FFF2CC",
    "Куртки": "F8CBAD",
    "двойка топ с завязками": "BDD7EE",
    "Песок": "FCE4D6",
    "Шанель": "E2EFDA",
    "Шорты": "D9E1F2",
    "Комбез": "F4B084",
    "Зара": "C6E0B4",
    "франц.лен": "E2EFDA",
    "платья": "F8CBAD"
    
}

# ====================== СОСТОЯНИЯ ======================

class Form(StatesGroup):
    waiting_for_file = State()

# ====================== СТАРТ ======================

@dp.message(Command("start"))
async def start_handler(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "Отправь список артикулов через запятую.\n"
        "После этого отправь Excel файл (.xlsx)"
    )

# ====================== СЛОВА ======================

@dp.message(F.text)
async def save_words(message: Message, state: FSMContext):
    words = [w.strip() for w in message.text.split(",") if w.strip()]
    await state.update_data(user_words=words)
    await state.set_state(Form.waiting_for_file)
    await message.answer("Слова сохранены. Теперь отправь Excel файл.")

# ====================== ФАЙЛ ======================

@dp.message(F.document, Form.waiting_for_file)
async def handle_document(message: Message, state: FSMContext):

    await message.answer("Файл получен. Обработка...")

    data = await state.get_data()
    user_words = data.get("user_words", [])

    file = await bot.get_file(message.document.file_id)
    downloaded = await bot.download_file(file.file_path)
    df = pd.read_excel(downloaded)

    col_article = 5   # F
    col_quantity = 34 # AI
    col_value = 36    # AK
    col_AJ = 35       # AJ

    results_by_group = {}
    avg_values = []
    total_sum = 0

    for group_name, group_list in groups.items():
        relevant = [word for word in group_list if word in user_words]
        if not relevant:
            continue

        results_by_group[group_name] = []

        for word in relevant:

            # обычный фильтр
            filtered = df[
                (df.iloc[:, col_article] == word) &
                (df.iloc[:, col_quantity] > 0)
            ]

            # фильтр возвратов
            filtered_returns = df[
                (df.iloc[:, col_article] == word) &
                (df.iloc[:, col_AJ] != 0) &
                (df.iloc[:, col_value] > 51)
            ]

            if filtered.empty:
                results_by_group[group_name].append([word, None, None, None, None, None, None])
            else:
                max_val = filtered.iloc[:, col_value].max()
                min_val = filtered.iloc[:, col_value].min()
                avg_val = filtered.iloc[:, col_value].mean()
                sum_val = filtered.iloc[:, col_quantity].sum()

                if filtered_returns.empty:
                    max_ret = None
                    min_ret = None
                else:
                    max_ret = filtered_returns.iloc[:, col_value].max()
                    min_ret = filtered_returns.iloc[:, col_value].min()

                results_by_group[group_name].append([
                    word, max_val, min_val, avg_val, sum_val, max_ret, min_ret
                ])

                avg_values.append(avg_val)
                total_sum += sum_val

    average_of_averages = sum(avg_values) / len(avg_values) if avg_values else 0

    # ====================== EXCEL ======================

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    yesterday = datetime.now() - timedelta(days=1)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Отчёт за {yesterday.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["Артикул", "Макс.", "Мин.", "Среднее", "Доставки (шт.)", "Макс", "Мин"])

    # Заголовок "Возвраты"
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.cell(row=2, column=6, value="Возвраты").font = Font(bold=True)
    ws.cell(row=2, column=6).alignment = Alignment(horizontal="center")

    for cell in ws[2]:
        cell.font = Font(bold=True)

    current_row = 3

    for group_name, rows in results_by_group.items():
        ws.cell(row=current_row, column=1, value=group_name).font = Font(bold=True, size=12)
        current_row += 1

        fill = PatternFill(start_color=group_colors[group_name], end_color=group_colors[group_name], fill_type="solid")

        for row_data in rows:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

    current_row += 1
    ws.cell(row=current_row, column=1, value="Доставок (шт.)").font = Font(bold=True)
    ws.cell(row=current_row, column=5, value=total_sum)

    current_row += 1
    ws.cell(row=current_row, column=1, value="Среднее от среднего").font = Font(bold=True)
    ws.cell(row=current_row, column=4, value=round(average_of_averages, 2))

    # автоширина
    for idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"отчёт_{yesterday.strftime('%d%m%Y')}.xlsx"
    await message.answer_document(
        BufferedInputFile(output.getvalue(), filename=file_name),
        caption="✅ Отчёт готов!"
    )

    await state.clear()

# ====================== ЗАПУСК ======================

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
